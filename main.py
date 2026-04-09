import atexit
import os
import random
import re
import sys
import time
from pathlib import Path
from datetime import datetime, date
from typing import Dict, List, Optional, TextIO, Tuple

import pandas as pd
import yaml
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import StaleElementReferenceException

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from langdetect import DetectorFactory, LangDetectException, detect

    DetectorFactory.seed = 0
except Exception:
    detect = None
    LangDetectException = Exception


# =========================
# Basic helpers
# =========================
PROJECT_DIR = Path(__file__).resolve().parent
PARAMS_PATH = PROJECT_DIR / "params.yaml"
LOGS_DIR = PROJECT_DIR / "logs"
_ACTIVE_LOG_FILE: Optional[TextIO] = None
_ORIGINAL_STDOUT = sys.stdout
_ORIGINAL_STDERR = sys.stderr


class TeeStream:
    def __init__(self, *streams: TextIO):
        self.streams = streams

    def write(self, data: str) -> int:
        for stream in self.streams:
            stream.write(data)
            stream.flush()
        return len(data)

    def flush(self) -> None:
        for stream in self.streams:
            stream.flush()

    def isatty(self) -> bool:
        return any(getattr(stream, "isatty", lambda: False)() for stream in self.streams)


def build_log_file_path(script_name: str) -> Path:
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    safe_name = re.sub(r"[^a-zA-Z0-9_-]+", "_", script_name).strip("_") or "run"
    return LOGS_DIR / f"{timestamp}_{safe_name}_log.txt"


def setup_run_logging(script_name: str) -> Path:
    global _ACTIVE_LOG_FILE

    if _ACTIVE_LOG_FILE is not None:
        return Path(_ACTIVE_LOG_FILE.name)

    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_file_path = build_log_file_path(script_name)
    log_file = open(log_file_path, "a", encoding="utf-8", buffering=1)

    started_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    header = [
        "=" * 80,
        f"Script started at: {started_at}",
        f"Script name: {script_name}",
        "=" * 80,
        "",
    ]
    log_file.write("\n".join(header))
    log_file.flush()

    sys.stdout = TeeStream(_ORIGINAL_STDOUT, log_file)
    sys.stderr = TeeStream(_ORIGINAL_STDERR, log_file)
    _ACTIVE_LOG_FILE = log_file
    return log_file_path


def close_run_logging() -> None:
    global _ACTIVE_LOG_FILE

    if _ACTIVE_LOG_FILE is None:
        return

    sys.stdout = _ORIGINAL_STDOUT
    sys.stderr = _ORIGINAL_STDERR
    _ACTIVE_LOG_FILE.close()
    _ACTIVE_LOG_FILE = None


atexit.register(close_run_logging)


def utc_now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")



def today_iso() -> str:
    return date.today().isoformat()



def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip().lower())



def extract_job_id_from_url(raw_url: str) -> Optional[str]:
    raw_url = (raw_url or "").strip()
    m = re.search(r"/jobs/view/(\d+)", raw_url)
    if not m:
        return None
    return m.group(1)



def canonical_job_url(raw_url: str) -> str:
    """
    Normalize LinkedIn job URLs so the same job doesn't create duplicates due to tracking params.
    Keeps only: https://www.linkedin.com/jobs/view/<job_id>/
    """
    job_id = extract_job_id_from_url(raw_url)
    if not job_id:
        return (raw_url or "").strip()
    return f"https://www.linkedin.com/jobs/view/{job_id}/"



def sleep_random(min_s: float, max_s: float) -> None:
    """Sleep random time between min_s and max_s seconds."""
    if min_s < 0:
        min_s = 0
    if max_s < min_s:
        max_s = min_s
    time.sleep(random.uniform(min_s, max_s))



def load_params(params_path) -> Dict:
    with open(params_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def resolve_job_names(params: Dict) -> List[str]:
    raw_job_names = params.get("job_name", "")

    if isinstance(raw_job_names, str):
        job_names = [raw_job_names]
    elif isinstance(raw_job_names, list):
        job_names = [str(item) for item in raw_job_names]
    else:
        raise SystemExit("Invalid 'job_name' in params.yaml. Use a string or a list of strings.")

    cleaned_job_names = [name.strip() for name in job_names if str(name).strip()]
    if not cleaned_job_names:
        raise SystemExit("No valid job names found in params.yaml. Fill 'job_name' with at least one value.")

    return cleaned_job_names



def build_linkedin_url(job_name: str, geo_id: int, remote_f_wt: Optional[int], sort_by_most_recent: Optional[int], start: int = 0) -> str:
    """
    Build a LinkedIn Jobs search URL using geoId and optional remote filter.
    Example:
      https://www.linkedin.com/jobs/search/?f_WT=2&geoId=92000000&keywords=Data%20Engineer&start=25
    """
    from urllib.parse import quote

    base = "https://www.linkedin.com/jobs/search/?"
    params = []
    if remote_f_wt is not None:
        params.append(f"f_WT={int(remote_f_wt)}")
    params.append(f"geoId={int(geo_id)}")
    params.append(f"keywords={quote(job_name)}")
    params.append(f"start={int(start)}")
    if sort_by_most_recent is not None:
        params.append("&sortBy=DD")
    return base + "&".join(params)+""



def init_driver(headless: bool) -> webdriver.Chrome:
    options = Options()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    return webdriver.Chrome(options=options)



def login_linkedin(driver: webdriver.Chrome, email: str, password: str, sleep_min: float, sleep_max: float) -> None:
    driver.get("https://www.linkedin.com/login")
    sleep_random(sleep_min, sleep_max)

    driver.find_element(By.ID, "username").send_keys(email)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    sleep_random(sleep_min, sleep_max)



def find_first_text_by_selectors(driver: webdriver.Chrome, selectors: List[str]) -> str:
    """Try multiple selectors and return the first non-empty text found."""
    for sel in selectors or []:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            text = (el.text or "").strip()
            if text:
                return text
        except Exception:
            continue
    return ""



def clean_job_title(raw_title: str) -> str:
    """
    Clean LinkedIn job titles that may include extra labels like 'with verification'.
    """
    t = (raw_title or "").strip()
    t = re.sub(r"\s+with verification\s*$", "", t, flags=re.IGNORECASE)
    return t.strip()


DESCRIPTION_PREFIX_PATTERNS = [
    r"^\s*sobre a vaga\b[:\s-]*",
    r"^\s*about the job\b[:\s-]*",
    r"^\s*job description\b[:\s-]*",
    r"^\s*descri(?:ç|c)[aã]o da vaga\b[:\s-]*",
]


def strip_description_prefix(raw_description: str) -> str:
    description = (raw_description or "").strip()
    if not description:
        return ""

    cleaned = description
    for pattern in DESCRIPTION_PREFIX_PATTERNS:
        cleaned = re.sub(pattern, "", cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def detect_text_language(text: str) -> Optional[str]:
    sample = strip_description_prefix(text)
    if not sample:
        return None

    # The detector gets more stable with a minimum amount of alphabetic text.
    alpha_only = re.sub(r"[^A-Za-zÀ-ÿ]+", " ", sample)
    if len(alpha_only.replace(" ", "")) < 30:
        return None

    if detect is None:
        return None

    try:
        return detect(sample)
    except LangDetectException:
        return None


def should_exclude_by_language(description: str, params: Dict) -> Optional[str]:
    if not bool(params.get("exclude_non_english_descriptions", False)):
        return None

    detected_language = detect_text_language(description)
    allowed_languages = {
        normalize_text(lang)
        for lang in (params.get("allowed_description_languages", ["en"]) or ["en"])
        if str(lang).strip()
    }

    if not detected_language:
        if bool(params.get("exclude_when_language_detection_fails", False)):
            return "language detection failed"
        return None

    if normalize_text(detected_language) not in allowed_languages:
        return f"description language: {detected_language}"

    return None


# =========================
# Keyword logic (ranking + filtering)
# =========================

def keyword_matches(text: str, weighted_keywords: Dict[str, int]) -> List[str]:
    """Return the list of keywords that appear in text (substring match after normalization)."""
    blob = normalize_text(text)
    matched = []
    for k in (weighted_keywords or {}).keys():
        if normalize_text(k) in blob:
            matched.append(k)
    return matched



def any_blocked(text: str, blocklist: List[str]) -> Optional[str]:
    """Return the blocking keyword if any is found, else None."""
    blob = normalize_text(text)
    for kw in blocklist or []:
        if normalize_text(kw) in blob:
            return kw
    return None


def matches_blocked_company(company: str, blocked_companies: List[str]) -> Optional[str]:
    company_norm = normalize_text(company)
    if not company_norm:
        return None

    for blocked_company in blocked_companies or []:
        blocked_norm = normalize_text(blocked_company)
        if blocked_norm and blocked_norm in company_norm:
            return blocked_company

    return None


def matches_location_filter(title: str, company: str, location: str, description: str, params: Dict) -> Optional[str]:
    if not bool(params.get("exclude_by_location_keywords", False)):
        return None

    location_keywords = params.get("location_filter_keywords", []) or []
    combined_text = " ".join([title, company, location, description]).strip()
    return find_matching_keyword(combined_text, location_keywords)


def should_exclude_job(title: str, company: str, location: str, description: str, params: Dict) -> Optional[str]:
    blocked_company = matches_blocked_company(company, params.get("blocked_companies", []) or [])
    if blocked_company:
        return f"blocked company: {blocked_company}"

    language_match = should_exclude_by_language(description, params)
    if language_match:
        return language_match

    location_match = matches_location_filter(title, company, location, description, params)
    if location_match:
        return f"location filter: {location_match}"

    return None


def exclusion_status_detail(excluded_by: str) -> str:
    return f"Skipped by company/location filter: {excluded_by}"


def summarize_job_for_log(title: str, company: str, location: str, url: str = "") -> str:
    parts = []
    if title.strip():
        parts.append(f"title='{title.strip()}'")
    if company.strip():
        parts.append(f"company='{company.strip()}'")
    if location.strip():
        parts.append(f"location='{location.strip()}'")
    if url.strip():
        parts.append(f"url='{url.strip()}'")
    return ", ".join(parts) if parts else "job metadata unavailable"



def compute_score(text: str, pos: Dict[str, int], neg: Dict[str, int]) -> Tuple[int, List[str], List[str]]:
    """
    score = sum(positive weights) - sum(negative weights)
    Returns: (score, matched_positive, matched_negative)
    """
    matched_pos = keyword_matches(text, pos)
    matched_neg = keyword_matches(text, neg)

    score_pos = sum(int(pos[k]) for k in matched_pos if k in pos)
    score_neg = sum(int(neg[k]) for k in matched_neg if k in neg)

    return (score_pos - score_neg), matched_pos, matched_neg


# =========================
# Output I/O
# =========================

OUTPUT_COLUMNS = [
    "title",
    "company",
    "location",
    "status",
    "linkedin_status",
    "score",
    "matched_positive_keywords",
    "matched_negative_keywords",
    "url",
    "job_id",
    "description",
    "description_language",
    "first_seen",
    "last_seen",
    "last_scraped_at",
    "status_detail",
    "linkedin_status_detail",
    "notes",
]


LOCAL_SKIP_NOTE_PATTERNS = [
    "not applying",
    "skip",
]


LOCAL_DERIVED_STATUSES = {"canceled", "cancelled", "skipped", "skiped"}


def ensure_output_schema(df: pd.DataFrame) -> pd.DataFrame:
    if "notes" not in df.columns and "Note" in df.columns:
        df["notes"] = df["Note"]

    for c in OUTPUT_COLUMNS:
        if c not in df.columns:
            df[c] = None

    missing_linkedin_status = df["linkedin_status"].isna() | (df["linkedin_status"].astype(str).str.strip() == "")
    current_status = df["status"].fillna("").astype(str)
    non_local_status = ~current_status.str.strip().str.lower().isin(LOCAL_DERIVED_STATUSES)
    df.loc[missing_linkedin_status & non_local_status, "linkedin_status"] = current_status

    missing_linkedin_detail = df["linkedin_status_detail"].isna() | (df["linkedin_status_detail"].astype(str).str.strip() == "")
    df.loc[missing_linkedin_detail, "linkedin_status_detail"] = df["status_detail"]

    df = df[OUTPUT_COLUMNS]
    df["url"] = df["url"].apply(lambda x: canonical_job_url(str(x)) if pd.notna(x) else "")
    df["job_id"] = df["url"].apply(lambda u: extract_job_id_from_url(u) or "")
    return df



def read_existing_output(output_file: str) -> pd.DataFrame:
    """Read existing output file if present, else return empty DataFrame with expected columns."""
    if not os.path.exists(output_file):
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    try:
        if output_file.lower().endswith(".csv"):
            df = pd.read_csv(output_file)
        else:
            df = pd.read_excel(output_file)
        return ensure_output_schema(df)
    except Exception:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)



def find_matching_keyword(text: str, keywords: List[str]) -> Optional[str]:
    blob = normalize_text(text)
    for kw in keywords or []:
        kw_str = str(kw).strip()
        kw_norm = normalize_text(kw_str)
        if kw_norm and kw_norm in blob:
            return kw_str
    return None


def derive_effective_row_state(row: pd.Series, params: Dict) -> Dict[str, object]:
    title = str(row.get("title") or "").strip()
    company = str(row.get("company") or "").strip()
    job_loc = str(row.get("location") or "").strip()
    description = str(row.get("description") or "").strip()
    description_language = detect_text_language(description) or ""
    notes = str(row.get("notes") or row.get("Note") or "").strip()

    combined_text = " ".join([title, company, job_loc, description]).strip()
    pos_keywords = params.get("positive_keywords", {}) or {}
    neg_keywords = params.get("negative_keywords", {}) or {}
    blocklist = params.get("blocklist_keywords", []) or []
    required_title_keywords = params.get("required_title_keywords", []) or []
    title_blocklist_keywords = params.get("title_blocklist_keywords", []) or []
    require_pos = bool(params.get("require_at_least_one_positive_keyword", True))
    allow_without_pos = bool(params.get("allow_add_without_positive_match", False))

    score, matched_pos, matched_neg = compute_score(combined_text, pos_keywords, neg_keywords)

    linkedin_status = str(row.get("linkedin_status") or "").strip()
    linkedin_detail = row.get("linkedin_status_detail")
    if not linkedin_status:
        current_status = str(row.get("status") or "").strip()
        if normalize_text(current_status) not in LOCAL_DERIVED_STATUSES:
            linkedin_status = current_status
            linkedin_detail = row.get("status_detail")

    effective_status = linkedin_status or "open"
    effective_detail = linkedin_detail
    title_norm = normalize_text(title)

    note_skip_match = find_matching_keyword(notes, LOCAL_SKIP_NOTE_PATTERNS)
    excluded_by = should_exclude_job(title, company, job_loc, description, params)
    title_blocked_by = find_matching_keyword(title, title_blocklist_keywords)
    blocked_by = any_blocked(combined_text, blocklist)
    missing_required_title = bool(required_title_keywords) and not any(
        normalize_text(k) in title_norm for k in required_title_keywords
    )
    lacks_required_positive = require_pos and (not matched_pos) and (not allow_without_pos)

    if note_skip_match:
        effective_status = "Skipped"
        effective_detail = f"Skipped by notes: {notes or note_skip_match}"
    elif excluded_by:
        effective_status = "Skipped"
        effective_detail = exclusion_status_detail(excluded_by)
    elif title_blocked_by:
        effective_status = "Canceled"
        effective_detail = f"Canceled by title_blocklist_keywords: {title_blocked_by}"
    elif blocked_by:
        effective_status = "Canceled"
        effective_detail = f"Canceled by blocklist_keywords: {blocked_by}"
    elif missing_required_title:
        effective_status = "Canceled"
        effective_detail = "Canceled by required_title_keywords"
    elif lacks_required_positive:
        effective_status = "Canceled"
        effective_detail = "Canceled by positive keyword policy"

    return {
        "score": int(score),
        "matched_positive_keywords": ", ".join(matched_pos),
        "matched_negative_keywords": ", ".join(matched_neg),
        "description_language": description_language,
        "linkedin_status": linkedin_status or "open",
        "linkedin_status_detail": linkedin_detail,
        "status": effective_status,
        "status_detail": effective_detail,
    }


def recalculate_output_rows(df: pd.DataFrame, params: Dict, emit_logs: bool = False) -> pd.DataFrame:
    df = ensure_output_schema(df.copy())

    for idx, row in df.iterrows():
        derived = derive_effective_row_state(row, params)
        for key, value in derived.items():
            df.at[idx, key] = value
        if emit_logs:
            excluded_by = should_exclude_job(
                str(row.get("title") or ""),
                str(row.get("company") or ""),
                str(row.get("location") or ""),
                str(row.get("description") or ""),
                params,
            )
            if excluded_by:
                print(
                    "[INFO] Existing output row marked as Skipped by company/location filter -> "
                    f"{excluded_by} | "
                    f"{summarize_job_for_log(str(row.get('title') or ''), str(row.get('company') or ''), str(row.get('location') or ''), str(row.get('url') or ''))}"
                )

    try:
        df["score"] = pd.to_numeric(df["score"], errors="coerce").fillna(0).astype(int)
        df = df.sort_values(by=["score", "last_seen"], ascending=[False, False]).reset_index(drop=True)
    except Exception:
        pass

    return df


def apply_status_formatting_xlsx(output_file: str, status_col_name: str = "status") -> None:
    wb = load_workbook(output_file)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    if status_col_name not in header:
        wb.save(output_file)
        return

    status_idx = header.index(status_col_name) + 1
    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row, column=status_idx).value
        status_norm = normalize_text(str(status_val)) if status_val is not None else ""

        fill = None
        if status_norm in {"closed", "canceled", "cancelled", "skipped", "skiped"}:
            fill = red_fill
        elif status_norm == "applied":
            fill = gray_fill

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
        else:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)

    wb.save(output_file)



def is_permission_denied_error(error: Exception) -> bool:
    msg = str(error or "").lower()
    return (
        isinstance(error, PermissionError)
        or "permission denied" in msg
        or "errno 13" in msg
        or "[errno 13]" in msg
    )


def build_failback_output_path(output_file: str) -> str:
    output_path = Path(output_file)
    return str(output_path.with_name("output_failback.xlsx"))


def write_output(df: pd.DataFrame, output_file: str, apply_formatting: bool = True) -> None:
    if output_file.lower().endswith(".csv"):
        df.to_csv(output_file, index=False)
        return

    df.to_excel(output_file, index=False)

    if apply_formatting:
        try:
            apply_status_formatting_xlsx(output_file, status_col_name="status")
        except Exception:
            pass


def write_output_with_failback(df: pd.DataFrame, output_file: str, apply_formatting: bool = True) -> Tuple[Optional[str], bool]:
    try:
        write_output(df, output_file, apply_formatting=apply_formatting)
        return output_file, False
    except Exception as e:
        if not is_permission_denied_error(e):
            raise

        failback_output = build_failback_output_path(output_file)
        print(
            f"[WARN] Incremental save permission denied for '{output_file}' -> {e}. "
            f"Trying failback save at '{failback_output}'."
        )

        try:
            write_output(df, failback_output, apply_formatting=apply_formatting)
            return failback_output, True
        except Exception as failback_error:
            print(f"[WARN] Failback save also failed for '{failback_output}' -> {failback_error}. Continuing.")
            return None, True


# =========================
# Job status detection
# =========================

def detect_job_status(driver, params: Dict) -> Tuple[str, Optional[str]]:
    """
    Detect job status based on UI messages in the job details page/panel.
    """
    selectors = params.get("status_selectors", {}) or {}
    closed_texts = params.get("status_closed_texts", []) or []
    applied_texts = params.get("status_applied_texts", []) or []

    applied_sel = (selectors.get("applied_banner") or "").strip()
    alert_sel = (selectors.get("any_alert_container") or "").strip()
    root_sel = (selectors.get("right_panel_root") or "body").strip()

    if applied_sel:
        try:
            el = driver.find_element(By.CSS_SELECTOR, applied_sel)
            msg = (el.text or "").strip()
            if msg:
                msg_norm = normalize_text(msg)
                for t in applied_texts:
                    if normalize_text(t) in msg_norm:
                        return "applied", msg
        except Exception:
            pass

    page_text = ""
    try:
        root = driver.find_element(By.CSS_SELECTOR, root_sel)
        page_text = (root.text or "").strip()
    except Exception:
        try:
            page_text = (driver.find_element(By.TAG_NAME, "body").text or "").strip()
        except Exception:
            page_text = ""

    page_text_norm = normalize_text(page_text)

    if "status da candidatura" in page_text_norm:
        for t in applied_texts:
            if normalize_text(t) in page_text_norm:
                return "applied", t

    if "application status" in page_text_norm:
        for t in applied_texts:
            if normalize_text(t) in page_text_norm:
                return "applied", t

    if "acessar site da empresa" in page_text_norm:
        for t in applied_texts:
            if normalize_text(t) in page_text_norm:
                return "applied", t
        return "applied", "Acessar site da empresa"

    if "visit company website" in page_text_norm or "go to company website" in page_text_norm:
        for t in applied_texts:
            if normalize_text(t) in page_text_norm:
                return "applied", t
        return "applied", "Company website application status"

    alert_text = ""
    if alert_sel:
        try:
            alerts = driver.find_elements(By.CSS_SELECTOR, alert_sel)
            alert_text = " ".join([(a.text or "").strip() for a in alerts if (a.text or "").strip()]).strip()
        except Exception:
            alert_text = ""

    blob = normalize_text(" ".join([page_text, alert_text]))

    for t in closed_texts:
        if normalize_text(t) in blob:
            return "closed", t

    for t in applied_texts:
        if normalize_text(t) in blob:
            return "applied", t

    return "open", None


# =========================
# Scrolling (left panel)
# =========================

def pick_scrollable_descendant(driver, root_el):
    try:
        candidate = driver.execute_script(
            """
            const root = arguments[0];
            const els = [root, ...root.querySelectorAll('*')];
            function isScrollable(el) {
              const style = window.getComputedStyle(el);
              const oy = style.overflowY;
              const canScroll = (oy === 'auto' || oy === 'scroll');
              return canScroll && el.scrollHeight > el.clientHeight + 5;
            }
            for (const el of els) {
              if (isScrollable(el)) return el;
            }
            return root;
            """,
            root_el,
        )
        return candidate
    except Exception:
        return root_el



def get_first_scroll_container(driver, selectors: List[str]):
    for sel in selectors or []:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            return sel, el
        except Exception:
            continue
    return None, None



def scroll_left_results_panel(driver,
                              container_selectors: List[str],
                              job_container_selector: str,
                              max_rounds: int,
                              pause_s: float,
                              sleep_min: float,
                              sleep_max: float) -> None:
    sel_used, root = get_first_scroll_container(driver, container_selectors)
    if not root:
        print(f"[WARN] Could not find left scroll container. Tried: {container_selectors}")
        return

    scroll_el = pick_scrollable_descendant(driver, root)
    print(f"[INFO] Using left scroll container selector: {sel_used}")

    prev_count = -1
    stable_rounds = 0

    for r in range(max_rounds):
        try:
            containers = driver.find_elements(By.CSS_SELECTOR, job_container_selector)
            count = len(containers)
        except Exception:
            containers = []
            count = 0

        if count == prev_count:
            stable_rounds += 1
        else:
            stable_rounds = 0

        if stable_rounds >= 3:
            break

        prev_count = count

        if containers:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'end'});", containers[-1])
            except Exception:
                pass

        try:
            driver.execute_script(
                "arguments[0].scrollTop = arguments[0].scrollTop + arguments[0].clientHeight;",
                scroll_el
            )
        except Exception:
            pass

        time.sleep(pause_s)
        sleep_random(sleep_min, sleep_max)

    print(f"[INFO] Left panel scroll completed (rounds={r+1})")


# =========================
# Scraping (stale-proof: snapshot URLs)
# =========================

def scrape_jobs(
    driver: webdriver.Chrome,
    params: Dict,
    existing_df: pd.DataFrame,
    output_file: str,
    job_name: str,
) -> pd.DataFrame:
    job_name = str(job_name).strip()
    max_pages = int(params.get("max_pages", 1))

    sleep_min = float(params.get("sleep_min_seconds", 1.0))
    sleep_max = float(params.get("sleep_max_seconds", 2.0))

    job_container_selector = str(params.get("job_container_selector", "")).strip()
    job_link_selector = str(params.get("job_link_selector", "")).strip()
    company_selector = str(params.get("company_selector", "")).strip()
    location_selector = str(params.get("location_selector", "")).strip()
    description_selectors = params.get("description_selectors", []) or []

    require_pos = bool(params.get("require_at_least_one_positive_keyword", True))
    allow_without_pos = bool(params.get("allow_add_without_positive_match", False))

    pos_keywords = params.get("positive_keywords", {}) or {}
    neg_keywords = params.get("negative_keywords", {}) or {}
    blocklist = params.get("blocklist_keywords", []) or []

    geo_id = int(params.get("geo_id", 92000000))
    remote_f_wt = params.get("remote_filter_f_wt", 2)
    sort_by_most_recent = bool(params.get("sort_by_most_recent", False))
    start_step = int(params.get("start_step", 25))

    ignored_job_ids = {str(x).strip() for x in (params.get("ignored_job_ids", []) or []) if str(x).strip()}
    save_after_each_job = bool(params.get("save_after_each_job", False))
    apply_row_formatting = bool(params.get("apply_row_formatting", True))
    required_title_keywords = params.get("required_title_keywords", []) or []
    title_blocklist_keywords = params.get("title_blocklist_keywords", []) or []

    if not job_container_selector or not job_link_selector:
        raise SystemExit("Missing required selectors in params.yaml: job_container_selector and job_link_selector")

    existing_by_url: Dict[str, int] = {}
    if not existing_df.empty:
        for idx, row in existing_df.iterrows():
            url = canonical_job_url(str(row.get("url") or "").strip())
            if url:
                existing_df.at[idx, "url"] = url
                existing_df.at[idx, "job_id"] = extract_job_id_from_url(url) or ""
                existing_by_url[url] = idx

    seen_in_this_run = set()

    print(f"[INFO] Starting search for job_name: {job_name}")

    for page in range(max_pages):
        start = page * start_step
        search_url = build_linkedin_url(job_name, geo_id, remote_f_wt, sort_by_most_recent, start=start)

        print(f"[INFO] [{job_name}] Page {page + 1}/{max_pages} -> Opening search URL: {search_url}")
        driver.get(search_url)
        sleep_random(sleep_min, sleep_max)

        scroll_left_results_panel(
            driver,
            container_selectors=params.get("left_list_scroll_container_selectors", []) or [],
            job_container_selector=job_container_selector,
            max_rounds=int(params.get("left_list_scroll_max_rounds", 30)),
            pause_s=float(params.get("left_list_scroll_pause_seconds", 1.0)),
            sleep_min=sleep_min,
            sleep_max=sleep_max,
        )

        try:
            containers = driver.find_elements(By.CSS_SELECTOR, job_container_selector)
        except Exception as e:
            print(f"[WARN] [{job_name}] Page {page + 1}: Failed to find containers -> {e}")
            containers = []

        job_urls: List[str] = []
        job_meta_by_url: Dict[str, Dict[str, str]] = {}

        for c in containers:
            try:
                link_el = c.find_element(By.CSS_SELECTOR, job_link_selector)
                url = canonical_job_url(link_el.get_attribute("href") or "")
                if not url:
                    continue

                job_id = extract_job_id_from_url(url) or ""
                if job_id and job_id in ignored_job_ids:
                    print(f"[INFO] [{job_name}] Page {page + 1}: Ignored by job_id during listing -> {job_id}")
                    continue

                title = (link_el.get_attribute("aria-label") or "").strip() or (link_el.text or "").strip()
                title = clean_job_title(title)
                title_norm = normalize_text(title)

                if required_title_keywords and not any(normalize_text(k) in title_norm for k in required_title_keywords):
                    print(
                        f"[INFO] [{job_name}] Page {page + 1}: Skipped due to title filter -> "
                        f"{summarize_job_for_log(title, '', '', url)}"
                    )
                    continue

                if title_blocklist_keywords and any(normalize_text(k) in title_norm for k in title_blocklist_keywords):
                    print(
                        f"[INFO] [{job_name}] Page {page + 1}: Skipped due to title blocklist -> "
                        f"{summarize_job_for_log(title, '', '', url)}"
                    )
                    continue

                company = ""
                if company_selector:
                    try:
                        company = (c.find_element(By.CSS_SELECTOR, company_selector).text or "").strip()
                    except Exception:
                        company = ""

                job_loc = ""
                if location_selector:
                    try:
                        job_loc = (c.find_element(By.CSS_SELECTOR, location_selector).text or "").strip()
                    except Exception:
                        job_loc = ""

                job_urls.append(url)
                job_meta_by_url[url] = {
                    "job_id": job_id,
                    "title": title,
                    "company": company,
                    "location": job_loc,
                }
            except Exception:
                continue

        seen = set()
        job_urls = [u for u in job_urls if not (u in seen or seen.add(u))]
        print(f"[INFO] [{job_name}] Page {page + 1}: Found {len(job_urls)} job URLs")

        for j, job_url in enumerate(job_urls, start=1):
            print(f"[INFO] [{job_name}] Page {page + 1}, Job {j} -> Processing URL...")
            sleep_random(sleep_min, sleep_max)

            try:
                if job_url in seen_in_this_run:
                    print(
                        f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Already processed in this run, skipping -> "
                        f"url='{job_url}'"
                    )
                    continue
                seen_in_this_run.add(job_url)

                job_id = extract_job_id_from_url(job_url) or ""
                if job_id and job_id in ignored_job_ids:
                    print(f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Ignored by job_id={job_id} -> url='{job_url}'")
                    continue

                clicked = False
                if job_id:
                    try:
                        fresh_link = driver.find_element(By.CSS_SELECTOR, f'a[href*="/jobs/view/"][href*="{job_id}"]')
                        driver.execute_script("arguments[0].click();", fresh_link)
                        clicked = True
                    except Exception:
                        clicked = False

                if not clicked:
                    driver.get(job_url)

                sleep_random(sleep_min, sleep_max)

                meta = job_meta_by_url.get(job_url, {})
                title = meta.get("title", "")
                company = meta.get("company", "")
                job_loc = meta.get("location", "")

                if not title:
                    try:
                        title = driver.find_element(By.CSS_SELECTOR, "h1").text.strip()
                        title = clean_job_title(title)
                    except Exception:
                        title = ""

                description = find_first_text_by_selectors(driver, description_selectors)
                status, status_detail = detect_job_status(driver, params)
                combined_text = " ".join([title, company, job_loc, description])

                excluded_by = should_exclude_job(title, company, job_loc, description, params)
                if excluded_by:
                    now_date = today_iso()
                    now_ts = utc_now_iso()
                    if job_url in existing_by_url:
                        idx = existing_by_url[job_url]
                        existing_df.at[idx, "last_seen"] = now_date
                        existing_df.at[idx, "last_scraped_at"] = now_ts
                        existing_df.at[idx, "job_id"] = job_id
                        if title:
                            existing_df.at[idx, "title"] = title
                        if company:
                            existing_df.at[idx, "company"] = company
                        if job_loc:
                            existing_df.at[idx, "location"] = job_loc
                        if description:
                            existing_df.at[idx, "description"] = description
                        existing_df.at[idx, "linkedin_status"] = status
                        existing_df.at[idx, "linkedin_status_detail"] = status_detail
                        existing_df.at[idx, "status"] = "Skipped"
                        existing_df.at[idx, "status_detail"] = exclusion_status_detail(excluded_by)
                        print(
                            f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Existing output row marked as Skipped -> "
                            f"{excluded_by} | {summarize_job_for_log(title, company, job_loc, job_url)}"
                        )
                        if save_after_each_job:
                            try:
                                saved_path, used_failback = write_output_with_failback(
                                    existing_df,
                                    output_file,
                                    apply_formatting=apply_row_formatting,
                                )
                                if used_failback and saved_path:
                                    print(
                                        f"[WARN] [{job_name}] Incremental save redirected to failback -> "
                                        f"{saved_path} (rows={len(existing_df)})"
                                    )
                                elif saved_path:
                                    print(f"[INFO] [{job_name}] Incremental save -> {saved_path} (rows={len(existing_df)})")
                                else:
                                    print(f"[WARN] [{job_name}] Incremental save failed and failback also failed. Continuing.")
                            except Exception as e:
                                print(f"[WARN] [{job_name}] Incremental save failed -> {e}")
                    else:
                        print(
                            f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Not added to output by company/location filter -> "
                            f"{excluded_by} | {summarize_job_for_log(title, company, job_loc, job_url)}"
                        )
                    continue

                blocked_by = any_blocked(combined_text, blocklist)
                if blocked_by:
                    print(
                        f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Not added to output by blocklist keyword -> "
                        f"{blocked_by} | {summarize_job_for_log(title, company, job_loc, job_url)}"
                    )
                    continue

                score, matched_pos, matched_neg = compute_score(combined_text, pos_keywords, neg_keywords)
                has_pos_match = len(matched_pos) > 0
                if require_pos and (not has_pos_match) and (not allow_without_pos):
                    print(
                        f"[INFO] [{job_name}] Page {page + 1}, Job {j}: Not added to output by positive keyword policy | "
                        f"{summarize_job_for_log(title, company, job_loc, job_url)}"
                    )
                    continue

                now_date = today_iso()
                now_ts = utc_now_iso()

                if job_url in existing_by_url:
                    idx = existing_by_url[job_url]
                    existing_df.at[idx, "last_seen"] = now_date
                    existing_df.at[idx, "last_scraped_at"] = now_ts

                    if title:
                        existing_df.at[idx, "title"] = title
                    if company:
                        existing_df.at[idx, "company"] = company
                    if job_loc:
                        existing_df.at[idx, "location"] = job_loc

                    existing_df.at[idx, "job_id"] = job_id
                    existing_df.at[idx, "description"] = description or existing_df.at[idx, "description"]
                    existing_df.at[idx, "linkedin_status"] = status
                    existing_df.at[idx, "linkedin_status_detail"] = status_detail
                    derived = derive_effective_row_state(existing_df.loc[idx], params)
                    for key, value in derived.items():
                        existing_df.at[idx, key] = value
                    print(f"[OK] [{job_name}] Page {page + 1}, Job {j}: Updated existing job")
                else:
                    new_row = {
                        "title": title,
                        "company": company,
                        "location": job_loc,
                        "status": status,
                        "linkedin_status": status,
                        "score": score,
                        "matched_positive_keywords": ", ".join(matched_pos),
                        "matched_negative_keywords": ", ".join(matched_neg),
                        "url": job_url,
                        "job_id": job_id,
                        "description": description,
                        "first_seen": now_date,
                        "last_seen": now_date,
                        "last_scraped_at": now_ts,
                        "status_detail": status_detail,
                        "linkedin_status_detail": status_detail,
                        "notes": "",
                    }
                    existing_df = pd.concat([existing_df, pd.DataFrame([new_row])], ignore_index=True)
                    existing_by_url[job_url] = existing_df.index[-1]
                    derived = derive_effective_row_state(existing_df.loc[existing_df.index[-1]], params)
                    for key, value in derived.items():
                        existing_df.at[existing_df.index[-1], key] = value
                    print(f"[OK] [{job_name}] Page {page + 1}, Job {j}: Added new job to output")

                if save_after_each_job:
                    try:
                        saved_path, used_failback = write_output_with_failback(
                            existing_df,
                            output_file,
                            apply_formatting=apply_row_formatting,
                        )
                        if used_failback and saved_path:
                            print(
                                f"[WARN] [{job_name}] Incremental save redirected to failback -> "
                                f"{saved_path} (rows={len(existing_df)})"
                            )
                        elif saved_path:
                            print(f"[INFO] [{job_name}] Incremental save -> {saved_path} (rows={len(existing_df)})")
                        else:
                            print(f"[WARN] [{job_name}] Incremental save failed and failback also failed. Continuing.")
                    except Exception as e:
                        print(f"[WARN] [{job_name}] Incremental save failed -> {e}")

            except StaleElementReferenceException as e:
                print(f"[WARN] [{job_name}] Page {page + 1}, Job {j}: Stale element -> {e}")
            except Exception as e:
                print(f"[WARN] [{job_name}] Page {page + 1}, Job {j}: Error processing job -> {e}")

        sleep_random(sleep_min, sleep_max)

    return recalculate_output_rows(existing_df, params)



def main():
    log_file_path = setup_run_logging("main")
    print(f"[INFO] Writing logs to: {log_file_path}")

    params = load_params(PARAMS_PATH)
    job_names = resolve_job_names(params)

    load_dotenv()
    email = os.getenv("LINKEDIN_EMAIL", "").strip()
    password = os.getenv("LINKEDIN_PASSWORD", "").strip()

    if not email or not password:
        raise SystemExit("Missing LINKEDIN_EMAIL / LINKEDIN_PASSWORD. Create a local .env file (not committed).")

    output_file = str(params.get("output_file", "jobs.xlsx")).strip()
    output_file = os.path.expanduser(output_file)
    output_file = os.path.abspath(output_file)
    headless = bool(params.get("headless", False))

    sleep_min = float(params.get("sleep_min_seconds", 1.0))
    sleep_max = float(params.get("sleep_max_seconds", 2.0))

    existing_df = read_existing_output(output_file)
    existing_df = recalculate_output_rows(existing_df, params, emit_logs=True)
    print(f"[INFO] Existing output rows: {len(existing_df)}")

    driver = init_driver(headless=headless)
    try:
        login_linkedin(driver, email, password, sleep_min, sleep_max)
        updated_df = existing_df
        total_job_names = len(job_names)

        for index, job_name in enumerate(job_names, start=1):
            print(f"[INFO] Starting batch search {index}/{total_job_names} for '{job_name}'")
            updated_df = scrape_jobs(driver, params, updated_df, output_file, job_name)

        saved_path, used_failback = write_output_with_failback(
            updated_df,
            output_file,
            apply_formatting=bool(params.get("apply_row_formatting", True)),
        )
        if used_failback and saved_path:
            print(f"[WARN] Final save redirected to failback -> {saved_path} (rows={len(updated_df)})")
        elif saved_path:
            print(f"[OK] Saved output file: {saved_path} (rows={len(updated_df)})")
        else:
            print("[WARN] Final save failed and failback also failed. Finishing without saving final file.")
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
